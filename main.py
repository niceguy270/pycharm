import urllib.request
import zipfile
import os
import xml.etree.ElementTree as ET
import openpyxl

crtfc_key = "23489435485ff33fc7141f5e01b6f54f676c1a0b"
corp_code = ""
stock_code = "005930"
bsns_year = "2022"
url = "https://opendart.fss.or.kr/api/corpCode.xml"
zip_name = "corpCode.zip"
xml_name = "CORPCODE.xml"

def getCorpCodeFile(url, crtfc_key):
    urllib.request.urlretrieve(url + "?crtfc_key=" + crtfc_key, zip_name)

    print(url + "?crfc_key=" + crtfc_key)
    print("파일 다운로드 완료")

    with zipfile.ZipFile(zip_name, 'r') as zip_ref:
        zip_ref.extractall()
        print("파일 압축해제 완료")
        zip_ref.close()
        os.remove(zip_name)

def getCorpCodebyStockCode(stock_code):
    tree = ET.parse(xml_name)
    root = tree.getroot()

    for child1 in root:
        corp_dic = {}

        for child2 in child1:
            print(child2.tag, child2.text)

            corp_dic[child2.tag] = child2.text

            if child2.tag == "stock_code":
                if stock_code == child2.text:
                    return corp_dic

wb = openpyxl.load_workbook("코스닥상장법인목록.xlsx")
ws = wb[wb.sheetnames[0]]

if getCorpCodeFile(url, crtfc_key) == True:


for i in range(1, ws.max_row):
    corp_dic = getCorpCodebyStockCode(ws.cell(row = i + 1, column = 2).value)
    print(corp_dic)