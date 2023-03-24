import openpyxl

wb = openpyxl.load_workbook("코스닥상장법인목록.xlsx")
ws = wb[wb.sheetnames[0]]

for i in range(1, ws.max_row):
    stock_code = ws.cell(row = i + 1, column = 2).value

